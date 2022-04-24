import json
from openpyxl import Workbook

# start by creating the 3 required sheets
workbook = Workbook()
sheet = workbook.active
sheet.title = 'survey'
workbook.create_sheet('choices')
workbook.create_sheet('settings')

# settings
with open('settings.json') as fh:
    settings = json.load(fh)
settings_sheet = workbook['settings']
for column, (key, value) in enumerate(settings.items()):
    if type(value) == list:
        value = ' '.join(value)
    settings_sheet.cell(row=1, column=column+1, value=key)
    settings_sheet.cell(row=2, column=column+1, value=value)

# choices
with open('choices.json') as fh:
    choices = json.load(fh)
choices_sheet = workbook['choices']
row = 1
choices_sheet.cell(row=row, column=1, value='list_name')
choices_sheet.cell(row=row, column=2, value='name')
choices_sheet.cell(row=row, column=3, value='label')
for key, options in choices.items():
    for choice, label in options.items():
        row += 1
        for i, value in enumerate([key, choice, label]):
            choices_sheet.cell(row=row, column=i+1, value=value)

# survey
with open('survey.json') as fh:
    survey = json.load(fh)
survey_sheet = workbook['survey']
columns = {}
next_column = 1
next_row = 2
for options in survey:
    for key, value in options.items():
        if type(value) == list:
            value = ' '.join(value)
        if key not in columns:
            columns[key] = next_column
            survey_sheet.cell(row=1, column=next_column, value=key)
            next_column += 1
        survey_sheet.cell(row=next_row, column=columns[key], value=value)
    next_row += 1

workbook.save(filename='C:/Users/marce/ArcGIS/My Survey Designs/Form 1/Form 1.xlsx')